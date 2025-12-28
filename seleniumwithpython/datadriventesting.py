from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time

# Load Excel
workbook = load_workbook('Book1.xlsx')
sheet = workbook.active

driver = webdriver.Chrome()
driver.maximize_window()
wait = WebDriverWait(driver, 10)

for row in sheet.iter_rows(min_row=2, values_only=True):
    username, password = row

    driver.get("https://www.saucedemo.com/")

    wait.until(EC.visibility_of_element_located((By.ID, "user-name"))).send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()

    try:
        # Wait for menu button → login success
        wait.until(EC.visibility_of_element_located((By.ID, "react-burger-menu-btn"))).click()
        wait.until(EC.visibility_of_element_located((By.ID, "logout_sidebar_link"))).click()
        print(f"Login successful: {username}")

    except:
        print(f"Login failed: {username}")

driver.quit()
